from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


OUT = Path("RAB_Gas_Test_Chamber.xlsx")

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
SUBHEADER_FILL = PatternFill("solid", fgColor="D9EAF7")
WARN_FILL = PatternFill("solid", fgColor="FCE4D6")
TOTAL_FILL = PatternFill("solid", fgColor="E2F0D9")
WHITE_FONT = Font(color="FFFFFF", bold=True)
BOLD = Font(bold=True)
SMALL = Font(size=9)
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def rupiah(value):
    if isinstance(value, (int, float)):
        return value
    return value


paket_a = [
    [1, "Box akrilik 30 x 30 x 30 cm", 1, "pcs", 243000, 243000, "Basis chamber kecil, perlu modifikasi port dan gasket"],
    [2, "Modifikasi port, gasket silikon, baut, dudukan sensor", 1, "lot", 150000, 150000, "Estimasi lokal/bengkel"],
    [3, "BME280 module", 1, "pcs", 58500, 58500, "Temperatur, kelembapan, tekanan"],
    [4, "Kipas brushless DC 12 V", 1, "pcs", 35000, 35000, "Mixing fan internal rendah tegangan"],
    [5, "Cable gland PG7", 5, "pcs", 2500, 12500, "Untuk kabel sensor/fan"],
    [6, "Selang PU pneumatic 6 mm", 5, "m", 6100, 30500, "Tubing gas/udara ringan"],
    [7, "Fitting pneumatic 6 mm", 10, "pcs", 5000, 50000, "Straight, elbow, tee, bulkhead"],
    [8, "Solenoid valve 12 V NC", 2, "pcs", 157500, 315000, "Gas inlet dan clean-air inlet; letakkan di luar chamber"],
    [9, "Regulator LPG SNI", 1, "pcs", 48138, 48138, "Untuk fase LPG-family sederhana"],
    [10, "Rotameter gas/udara sederhana", 1, "pcs", 190000, 190000, "Minimal flow indication, bukan mass flow controller"],
    [11, "Portable combustible gas leak detector / LEL detector low-cost", 1, "pcs", 750000, 750000, "Safety monitor eksternal, bukan reference-grade"],
    [12, "Power supply 12 V DC", 1, "pcs", 100000, 100000, "Fan + solenoid"],
    [13, "Clamp, seal tape, bracket, kabel, konektor, terminal", 1, "lot", 150000, 150000, "Consumables"],
]

paket_b = [
    [1, "Chamber akrilik custom 10-30 L dengan port rapi", 1, "pcs", 750000, 507000, "Upgrade dari box akrilik sederhana"],
    [2, "Gas rotameter LZM-6T 15-20 L/min", 1, "pcs", 1049000, 859000, "Lebih sesuai untuk gas dibanding rotameter murah"],
    [3, "4-in-1 gas detector CO/H2S/O2/LEL", 1, "pcs", 2250000, 1500000, "Replace detector murah; tetap perlu kalibrasi"],
    [4, "BME280/SHT module kualitas lebih baik", 1, "pcs", 148400, 89900, "Upgrade dari modul murah"],
    [5, "Exhaust fan/ducting kecil + bracket", 1, "lot", 300000, 300000, "Venting ke lokasi aman"],
    [6, "Extra fittings, check valve, quick connector", 1, "lot", 250000, 250000, "Membuat flow path lebih mudah dirawat"],
]

paket_c = [
    [1, "4-in-1 detector CO/H2S/O2/LEL low-mid range", 1, "pcs", "Rp2.199.000 - Rp3.600.000", "", "Minimum safety monitor, perlu cek kalibrasi"],
    [2, "4-gas detector brand industrial higher-end", 1, "pcs", 13320000, 13320000, "Contoh BW Max XT II, lebih dekat ke industrial"],
    [3, "CO meter electrochemical portable", 1, "pcs", "Rp495.000+", "", "Untuk CO only; cek sertifikat/kalibrasi"],
    [4, "Certified calibration gas / jasa kalibrasi", 1, "lot", "RFQ", "", "Wajib untuk reference-grade"],
    [5, "Exhaust/fume handling dan safety procedure", 1, "lot", "RFQ", "", "Wajib untuk CO/H2S"],
]

reference_options = [
    ["Butane / LPG", "NevadaNano MPS flammable gas sensor atau calibrated LPG/LEL detector", "TGS2610 hanya cross-check"],
    ["Propane", "NevadaNano MPS flammable gas sensor atau calibrated LPG/LEL detector", "TGS2610 hanya cross-check"],
    ["Methane", "NevadaNano MPS flammable gas sensor atau calibrated methane/LEL detector", "Jangan pakai TGS2610 sebagai ppm ground truth"],
    ["CO", "Alphasense CO electrochemical sensor atau calibrated CO detector/logger", "Butuh safety dan kalibrasi"],
    ["H2S", "Alphasense H2S electrochemical sensor atau calibrated H2S detector/logger", "Butuh safety dan kalibrasi"],
    ["VOC / false positives", "Sensirion SGP40 atau sejenis", "Interference context only, bukan ppm ground truth"],
]

sources = [
    ["Box akrilik 30 x 30 x 30 cm", "Blibli", "Rp243.000", "https://www.blibli.com/p/kotak-amal-akrilik-ukuran-besar-30x30x30-cm/ps--DUC-60031-00363"],
    ["BME280 murah", "Blibli", "Rp58.500", "https://www.blibli.com/p/gy-bme280-bme-280-sensor-temperatur-kelembapan-barometric-tekanan-iic/ps--ITE-70030-00189"],
    ["BME280 lebih baik", "Digiware", "Rp148.400", "https://digiwarestore.com/en/sensor/bme280-environmental-sensor-temperature-humidity-barometric-pressure-296523.html"],
    ["Kipas brushless 12 V", "Blibli", "Rp10.000 - Rp125.000", "https://www.blibli.com/jual/dc-blower-fan-12-v"],
    ["Cable gland PG7", "Blibli/Lazada", "Rp2.000 - Rp5.660", "https://www.blibli.com/jual/gland-kabel-pg7-pvc"],
    ["Selang PU 6 mm", "BigGo/Shopee aggregate", "Rp5.000 - Rp6.200/m", "https://biggo.id/s/pneumatic%20selang%206mm"],
    ["Fitting pneumatic 6 mm", "Blibli/Shopee aggregate", "Rp2.500 - Rp9.500/pcs", "https://biggo.id/s/pneumatic%20fitting%206mm"],
    ["Solenoid valve 12 V", "BigGo/Shopee aggregate", "Sekitar Rp157.500", "https://biggo.id/s/Solenoid%2BValve%2B12V"],
    ["Regulator LPG SNI", "Shopee", "Rp48.138", "https://shopee.co.id/Regulator-Gas-LPG-Elpiji-SNI-Termurah-dan-Terbaik-i.641642292.15242832864"],
    ["Rotameter search", "Shopee", "Sekitar Rp190.000+", "https://shopee.co.id/search?keyword=rotameter+flow+meter"],
    ["Gas rotameter LZM-6T", "Shopee", "Rp1.049.000", "https://shopee.co.id/Dn12-1-4-Inch-Flow-15-L-Min-Lzm-6T-Rotameter-%28Gas%29-i.164601711.27812346631"],
    ["Portable combustible detector", "Shopee", "Rp547.000 - Rp1.150.000", "https://shopee.co.id/search?keyword=gas+detectors"],
    ["4-in-1 detector", "SNDWAY Indonesia", "Rp2.250.000", "https://www.sndway.id/product-page/sndway-gas-detector-4-in-1-co-h2s-o2-combustible-gas-sensor-sw-7500a-pro"],
    ["NevadaNano MPS", "Official", "%LEL flammable gases", "https://nevadanano.com/products/mps-flammable-gas-sensor/"],
    ["Alphasense CO", "Official", "CO electrochemical sensors", "https://www.alphasense.com/products/view-by-target-gas/carbon-monoxide-sensors-co"],
    ["Alphasense H2S", "Official", "H2S electrochemical sensors", "https://www.alphasense.com/products/view-by-target-gas/hydrogen-sulphide-sensors-h2s"],
    ["Sensirion SGP40", "Official", "VOC index/interference context", "https://sensirion.com/sgp40"],
]


def style_range(ws):
    for row in ws.iter_rows():
        for cell in row:
            cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if isinstance(cell.value, (int, float)) and cell.column in (5, 6):
                cell.number_format = '"Rp"#,##0'


def set_widths(ws, widths):
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width


def add_title(ws, title, subtitle=None):
    ws["A1"] = title
    ws["A1"].font = Font(size=16, bold=True, color="1F4E78")
    if subtitle:
        ws["A2"] = subtitle
        ws["A2"].font = Font(size=10, italic=True, color="666666")


def add_budget_sheet(wb, name, title, data, total_formula=True, notes=None):
    ws = wb.create_sheet(name)
    add_title(ws, title, "Harga estimasi riset 2026-04-30. Belum termasuk ongkir, pajak, diskon, atau stok habis.")
    headers = ["No", "Item", "Qty", "Satuan", "Harga Satuan", "Total / Tambahan", "Catatan"]
    ws.append([])
    ws.append(headers)
    header_row = 4
    for cell in ws[header_row]:
        cell.fill = HEADER_FILL
        cell.font = WHITE_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in data:
        ws.append(row)
    last = ws.max_row
    if total_formula:
        ws.append(["", "", "", "", "Subtotal", f"=SUM(F5:F{last})", ""])
        subtotal_row = ws.max_row
        ws.append(["", "", "", "", "Cadangan 15%", f"=F{subtotal_row}*15%", ""])
        reserve_row = ws.max_row
        ws.append(["", "", "", "", "Estimasi Total", f"=F{subtotal_row}+F{reserve_row}", ""])
        for row_idx in [subtotal_row, reserve_row, ws.max_row]:
            for cell in ws[row_idx]:
                cell.fill = TOTAL_FILL
                cell.font = BOLD
    if notes:
        ws.append([])
        ws.append(["Catatan", notes])
        ws[ws.max_row][0].fill = WARN_FILL
        ws[ws.max_row][0].font = BOLD
    set_widths(ws, [6, 42, 8, 10, 18, 18, 45])
    style_range(ws)
    ws.freeze_panes = "A5"
    return ws


wb = Workbook()
ws = wb.active
ws.title = "Ringkasan"
add_title(ws, "RAB Gas Test Chamber", "Disusun dari riset harga marketplace Indonesia dan rekomendasi hardware chamber gas.")

summary_rows = [
    ["Paket A", "Starter LPG-family chamber", 2452534, "Validasi mekanik, logging awal, belum untuk klaim ppm produksi"],
    ["Paket B", "Improved controlled-flow LPG/methane chamber", 5958434, "Lebih repeatable, ada detector 4-in-1 low-mid range"],
    ["Paket C", "Toxic gas / reference-grade upgrade", "RFQ / Rp2,2 juta - Rp13,3 juta+", "Untuk CO/H2S hanya dengan safety lab dan kalibrasi"],
]
ws.append([])
ws.append(["Paket", "Deskripsi", "Estimasi Total", "Catatan"])
for cell in ws[4]:
    cell.fill = HEADER_FILL
    cell.font = WHITE_FONT
for row in summary_rows:
    ws.append(row)

ws.append([])
ws.append(["Peringatan keselamatan", "CO dan H2S tidak direkomendasikan untuk chamber DIY tanpa instrumen tersertifikasi, ventilasi, prosedur darurat, dan supervisi lab."])
ws[ws.max_row][0].fill = WARN_FILL
ws[ws.max_row][0].font = BOLD
ws[ws.max_row][1].fill = WARN_FILL

set_widths(ws, [16, 42, 22, 70])
style_range(ws)
for row in range(5, 7):
    ws.cell(row, 3).number_format = '"Rp"#,##0'

add_budget_sheet(wb, "Paket A", "Paket A - Starter LPG-Family Chamber", paket_a)
add_budget_sheet(wb, "Paket B", "Paket B - Improved Controlled-Flow LPG/Methane Chamber", paket_b)
add_budget_sheet(
    wb,
    "Paket C",
    "Paket C - Toxic Gas / Reference-Grade Upgrade",
    paket_c,
    total_formula=False,
    notes="Gunakan hanya dengan instrumen tersertifikasi, ventilasi, SOP darurat, dan supervisi lab.",
)

ws = wb.create_sheet("Reference Sensors")
add_title(ws, "Rekomendasi Reference Sensors", "MQ array tetap menjadi input ML; reference sensor/instrument dipakai untuk label dan validasi.")
ws.append([])
ws.append(["Target Gas", "Recommended Reference Option", "Catatan"])
for cell in ws[4]:
    cell.fill = HEADER_FILL
    cell.font = WHITE_FONT
for row in reference_options:
    ws.append(row)
set_widths(ws, [22, 70, 55])
style_range(ws)

ws = wb.create_sheet("Sumber Harga")
add_title(ws, "Sumber Harga Dan Link", "Prioritas Shopee/Tokopedia. Beberapa item memakai Blibli, Digiware, BigGo aggregate, atau official source.")
ws.append([])
ws.append(["Kebutuhan", "Sumber", "Harga / Info", "Link"])
for cell in ws[4]:
    cell.fill = HEADER_FILL
    cell.font = WHITE_FONT
for row in sources:
    ws.append(row)
    link_cell = ws.cell(ws.max_row, 4)
    if str(link_cell.value).startswith("http"):
        link_cell.hyperlink = link_cell.value
        link_cell.style = "Hyperlink"
set_widths(ws, [34, 24, 28, 90])
style_range(ws)

ws = wb.create_sheet("Urutan Beli")
add_title(ws, "Urutan Pembelian Yang Disarankan")
items = [
    ["1", "Box/chamber akrilik kecil atau custom chamber 10-30 L"],
    ["2", "BME280/SHT sensor"],
    ["3", "Fan brushless 12 V"],
    ["4", "Cable gland, gasket, fittings, tubing"],
    ["5", "Dua solenoid 12 V normally-closed"],
    ["6", "Regulator dan rotameter"],
    ["7", "Portable combustible gas detector / LEL detector"],
    ["8", "Setelah chamber stabil: upgrade ke 4-in-1 gas detector atau reference instrument"],
]
ws.append([])
ws.append(["Prioritas", "Item"])
for cell in ws[4]:
    cell.fill = HEADER_FILL
    cell.font = WHITE_FONT
for row in items:
    ws.append(row)
ws.append([])
ws.append(["Jangan diprioritaskan dulu", "Raw H2S sensor tanpa calibration setup; raw CO sensor untuk klaim ppm; mass flow controller mahal sebelum chamber dasar terbukti; CO/H2S testing sebelum SOP lab."])
ws[ws.max_row][0].fill = WARN_FILL
ws[ws.max_row][0].font = BOLD
ws[ws.max_row][1].fill = WARN_FILL
set_widths(ws, [18, 95])
style_range(ws)

for sheet in wb.worksheets:
    sheet.sheet_view.showGridLines = False

wb.save(OUT)
print(f"Saved {OUT}")
